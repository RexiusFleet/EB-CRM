#!/usr/bin/env python3
"""
ETL part 2: pricing workbook -> rate card, engine constants, quote history,
blow-time + drive-time benchmarks.

Input : Eugene Blower Pricing 2026 v2.xlsm
Output: data/rate_card.csv, data/markup_curve.csv, data/settings.csv,
        data/quote_history.csv, data/blow_benchmarks.csv, data/drive_zones.csv
"""
import openpyxl, csv, os, re, json, statistics, collections, sys

SRC = sys.argv[1] if len(sys.argv) > 1 else (
    '/root/.claude/uploads/12cbdb16-281c-59b8-8a9d-1b95b84ac0c8/'
    '89cfa59d-Eugene_Blower_Pricing_2026_v2.xlsm')
HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, 'data')
os.makedirs(OUT, exist_ok=True)
sys.path.insert(0, HERE)
from addr import street_key, extract_city, extract_zip
from contacts import clean_email, clean_phone


def pct(sorted_vals, p):
    if not sorted_vals:
        return None
    k = max(0, min(len(sorted_vals) - 1, int(round((len(sorted_vals) - 1) * p))))
    return round(sorted_vals[k], 4)


def band(y):
    if y is None:
        return None
    for lo, hi, name in [(0, 10, '1-10'), (10, 20, '11-20'), (20, 35, '21-35'),
                         (35, 60, '36-60'), (60, 1e9, '60+')]:
        if lo < y <= hi:
            return name
    return None

YARDS_PER_UNIT = 7.5
wb = openpyxl.load_workbook(SRC, data_only=True)

# ---------------------------------------------------------------- rate card
pc = wb['Product Costs']
rate_card = []
for r in range(3, 36):
    name = pc.cell(r, 2).value
    if not name or not str(name).strip():
        continue
    name = str(name).strip()
    unit_cost = pc.cell(r, 5).value
    cat = {'B': 'bark', 'S': 'soil_compost', 'R': 'rock'}.get(
        str(pc.cell(r, 17).value or '').strip().upper(), 'other')
    has = isinstance(unit_cost, (int, float)) and unit_cost > 0
    rate_card.append({
        'product': name,
        'erp_code': str(pc.cell(r, 1).value or '').strip() or None,
        'category': cat,
        'unit_cost': round(unit_cost, 4) if has else None,
        'yards_per_unit': YARDS_PER_UNIT,
        'cost_per_yard': round(unit_cost / YARDS_PER_UNIT, 4) if has else None,
        'needs_cost': not has,
        'active': True,
        'source': 'Eugene Blower Pricing 2026 v2.xlsm / Product Costs',
    })

# ---------------------------------------------------------------- markup curve
pw = wb['Pricing Worksheet']
markup = []
for c in range(3, 50):
    y, m = pw.cell(8, c).value, pw.cell(7, c).value
    if isinstance(y, (int, float)) and isinstance(m, (int, float)):
        markup.append({'min_yards': y, 'markup': round(m, 4)})
markup.sort(key=lambda x: x['min_yards'])
# collapse to bands (only keep rows where the markup actually changes)
curve, last = [], None
for row in markup:
    if row['markup'] != last:
        curve.append(row)
        last = row['markup']

# ---------------------------------------------------------------- settings
qc = wb['Quote Calculator JD']
settings = [
    ('labor_rate_blended', 27.50, '$/hr; avg(driver 30, helper 25) - Quote Calculator K14'),
    ('driver_wage', 30.0, '$/hr - Quote Calculator I13'),
    ('helper_wage', 25.0, '$/hr - Quote Calculator I19'),
    ('prevailing_wage', 70.0, '$/hr; replaces both wages when PW job - I12/I18'),
    ('equipment_rate', 85.0, '$/hr - Quote Calculator I22'),
    ('overhead_per_yard', 25.0, '$/yd, applied when yards < threshold - I23'),
    ('overhead_per_equip_hour', 145.0, '$/equip-hr, applied at/above threshold - I24'),
    ('overhead_yard_threshold', 16.0, 'yards; below = per-yard OH, at/above = hourly OH'),
    ('prep_hours_default', 0.5, 'hr; constant on 1974/2024 saved quotes'),
    ('helpers_default', 1.0, 'crew helpers; 1 on 1870/2024 saved quotes'),
    ('yards_per_unit', 7.5, 'yards per truck unit'),
]
settings = [{'key': k, 'value': v, 'note': n} for k, v, n in settings]

# ---------------------------------------------------------------- quote history
sq = wb['Saved Quotes']
HX = {sq.cell(1, c).value: c for c in range(1, sq.max_column + 1)}


def qg(r, key):
    v = sq.cell(r, HX[key]).value if key in HX else None
    if isinstance(v, str):
        v = v.strip()
        if v in ('#REF!', '#N/A', '#VALUE!', ''):
            return None
    if v == 0 and key in ('Name', 'Phone Number', 'Email Address', 'Address', 'Product '):
        return None
    return v


def numf(v):
    return round(float(v), 4) if isinstance(v, (int, float)) else None


def split_addr(a):
    """'740 Queens Ave, Creswell, OR 97426' -> (street_key, city, zip)"""
    if not a:
        return None, None, None
    return street_key(a) or None, extract_city(a), extract_zip(a)


quotes, qrecs = [], []
for r in range(2, sq.max_row + 1):
    prod, qty = qg(r, 'Product '), qg(r, 'Quantity')
    d = qg(r, 'Date')
    if not isinstance(prod, str) or not isinstance(qty, (int, float)) or not qty:
        continue
    if not hasattr(d, 'year'):
        continue
    skey, city, zc = split_addr(qg(r, 'Address'))
    prep, drive, blow, helpers = (numf(qg(r, k)) for k in ('Prep', 'Drive', 'Blow', 'Help'))
    rec = {
        'source_row': r,
        'quote_date': d.date().isoformat(),
        'sales_person': qg(r, 'Sales Person'),
        'customer_name': qg(r, 'Name'),
        'phone': clean_phone(qg(r, 'Phone Number')),
        'email': clean_email(qg(r, 'Email Address')),
        'address_raw': qg(r, 'Address'),
        'city': city, 'zip': zc,
        'site_key': skey,
        'product': prod.strip(),
        'yards': numf(qty),
        'prep_hours': prep, 'drive_hours': drive,
        'blow_hours': blow, 'helpers': helpers,
        'labor_hours': numf(qg(r, 'Labor Hours')),
        'equip_hours': numf(qg(r, 'Equip Hours')),
        'projected_cost': numf(qg(r, 'Projected Cost')),
        'target_price': numf(qg(r, 'Target Price')),
        'bid_amount': numf(qg(r, 'Bid Amount')),
        'material_rev': numf(qg(r, 'Material Rev')),
        'labor_rev': numf(qg(r, 'Labor Rev')),
        'equip_rev': numf(qg(r, 'Equipment Rev')),
        'notes': qg(r, 'Notes'),
    }
    quotes.append(rec)
    if all(isinstance(x, (int, float)) for x in (prep, drive, blow)):
        qrecs.append(rec)

# ---------------------------------------------- reconcile with ERP sites
# The ERP records "33860 Oak Springs Dr"; the quote log records
# "33860 Oak Springs Coburg". After the city is stripped those are
# "33860 oak springs dr" vs "33860 oak springs" -- one property, two keys, so
# the delivered history never appeared against the quote. Restore the suffix
# from the ERP, but ONLY where exactly one ERP street matches; genuinely
# ambiguous cases ("1767 Walnut" = Dr and St) are left split rather than guessed.
SUFFIX_WORDS = {'ave', 'st', 'rd', 'dr', 'ln', 'ct', 'blvd', 'pl', 'ter', 'cir',
                'pkwy', 'hwy', 'way', 'loop', 'trl', 'sq', 'xing', 'hts', 'rdg',
                'vw', 'mt'}


def _stem(k):
    t = (k or '').split()
    return ' '.join(t[:-1]) if len(t) > 1 and t[-1] in SUFFIX_WORDS else k


_erp_keys = {r['site_key'] for r in csv.DictReader(
    open(os.path.join(OUT, 'orders.csv'))) if r['site_key']}
_by_stem = collections.defaultdict(set)
for _k in _erp_keys:
    _by_stem[_stem(_k)].add(_k)

_merged = 0
for rec in quotes:
    k = rec['site_key']
    if k and k not in _erp_keys:
        cands = _by_stem.get(k, set())
        if len(cands) == 1:
            rec['site_key'] = next(iter(cands))
            _merged += 1

# The quote log also glues the city onto the address ("2096 Musket St Eugene"),
# which then displays twice beside the city column.
def _tidy_address(a, city):
    if not a:
        return a
    s = re.sub(r'[\s,]+\d{5}(-\d{4})?$', '', str(a).strip())
    s = re.sub(r'[\s,]+(OR|ORE|OREGON|WA|CA)\.?$', '', s, flags=re.I)
    if city:
        s = re.sub(r'[\s,]+' + re.escape(city) + r'\.?$', '', s, flags=re.I)
    return s.strip().rstrip(',') or None


for rec in quotes:
    rec['address_raw'] = _tidy_address(rec['address_raw'], rec['city'])

# ------------------------------------------------- blow-time benchmarks
# SOURCE: the ERP, not the quote log.
#
# The ERP has no explicit "blow time" column -- it records BILLED equipment
# hours (the Retail-Equipment Hours line). Those turn out to be an excellent
# stand-in: across the 18 products present in both sources, the median ratio of
# ERP equipment-hours-per-yard to quoted blow-hours-per-yard is 1.019. So the
# blower truck's billed time is, in practice, the blow time.
#
# Two reasons this is the better source:
#   - it is what actually happened and was invoiced, not what someone estimated
#   - 4,249 observations against 2,247 in the quote log
#
# ERP product descriptions are free text, so they are canonicalised first
# (see products.py) or a single product splinters across dozens of spellings.
erp_orders = list(csv.DictReader(open(os.path.join(OUT, 'orders.csv'))))


def _f(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


bb = collections.defaultdict(list)
for r in erp_orders:
    if r.get('order_type') != 'blow_in':
        continue
    prod = r.get('product')            # canonical name, from products.py
    y, e = _f(r.get('total_yards')), _f(r.get('equip_hours'))
    if not (prod and y and e):
        continue
    hy = e / y
    if not (0 < hy < 3):               # drop data-entry outliers
        continue
    bb[(prod, band(y) or 'ALL')].append(hy)
    bb[(prod, 'ALL')].append(hy)

# Quote-derived figures are kept only as a fallback for products the ERP has
# too little of, so a thin product still gets an estimate rather than the
# generic 0.10 hr/yd.
qfall = collections.defaultdict(list)
for q in qrecs:
    if q['yards'] and q['blow_hours'] is not None and q['yards'] > 0:
        hy = q['blow_hours'] / q['yards']
        if 0 < hy < 3:
            qfall[(q['product'], band(q['yards']) or 'ALL')].append(hy)
            qfall[(q['product'], 'ALL')].append(hy)

MIN_N = 5
blow_rows = []
seen = set()
for (prod, b), vals in sorted(bb.items()):
    sv = sorted(vals)
    if len(sv) < 3:
        continue
    seen.add((prod, b))
    blow_rows.append({
        'product': prod, 'volume_band': b, 'n': len(sv),
        'blow_hr_per_yard_p25': pct(sv, .25),
        'blow_hr_per_yard_median': pct(sv, .50),
        'blow_hr_per_yard_p75': pct(sv, .75),
        'source': 'erp_equipment_hours',
    })
for (prod, b), vals in sorted(qfall.items()):
    if (prod, b) in seen:
        continue
    sv = sorted(vals)
    if len(sv) < MIN_N:
        continue
    blow_rows.append({
        'product': prod, 'volume_band': b, 'n': len(sv),
        'blow_hr_per_yard_p25': pct(sv, .25),
        'blow_hr_per_yard_median': pct(sv, .50),
        'blow_hr_per_yard_p75': pct(sv, .75),
        'source': 'quote_blow_hours',
    })

# ------------------------------------------------- drive-time zones
dz = collections.defaultdict(list)
for q in qrecs:
    if q['drive_hours'] is None or not (0 < q['drive_hours'] <= 8):
        continue
    if q['city']:
        dz[('city', q['city'])].append(q['drive_hours'])
    if q['zip']:
        dz[('zip', q['zip'])].append(q['drive_hours'])

drive_rows = []
for (kind, key), vals in sorted(dz.items()):
    s = sorted(vals)
    if len(s) < 3:
        continue
    drive_rows.append({
        'zone_type': kind, 'zone_key': key, 'n': len(s),
        'drive_hours_p25': pct(s, .25),
        'drive_hours_median': pct(s, .50),
        'drive_hours_p75': pct(s, .75),
    })
allv = sorted(q['drive_hours'] for q in qrecs
              if q['drive_hours'] and 0 < q['drive_hours'] <= 8)
drive_rows.append({'zone_type': 'default', 'zone_key': '*', 'n': len(allv),
                   'drive_hours_p25': pct(allv, .25),
                   'drive_hours_median': pct(allv, .50),
                   'drive_hours_p75': pct(allv, .75)})

# ---------------------------------------------------------------- write


def dump(name, recs):
    if not recs:
        print(f'  {name:24} EMPTY'); return
    with open(os.path.join(OUT, name), 'w', newline='') as f:
        w = csv.DictWriter(f, fieldnames=list(recs[0].keys()))
        w.writeheader(); w.writerows(recs)
    print(f'  {name:24} {len(recs):>6} rows')


print('Writing pricing CSVs:')
dump('rate_card.csv', rate_card)
dump('markup_curve.csv', curve)
dump('settings.csv', settings)
dump('quote_history.csv', quotes)
dump('blow_benchmarks.csv', blow_rows)
dump('drive_zones.csv', drive_rows)

need = [r['product'] for r in rate_card if r['needs_cost']]
print(f'\nProducts flagged needs_cost ({len(need)}): {", ".join(need)}')
print(f'Markup curve: ' + ', '.join(f"{c['min_yards']}yd+ -> {c['markup']:.0%}" for c in curve))
matched = sum(1 for q in quotes if q['site_key'])
print(f'Quote history: {len(quotes)} rows, {matched} with a resolvable site_key')
print(f'Drive zones: {len(drive_rows)} (incl. global default '
      f'{drive_rows[-1]["drive_hours_median"]} hr)')
